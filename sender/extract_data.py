import json
import os

import openpyxl as openpyxl
import pandas as pd
from math import ceil, isnan

from Entities.SigfoxProfile import SigfoxProfile


def extract_data(foldername, stats_file, output, spreadsheet):
    pd.set_option('display.max_columns', None)
    path = f"{foldername}/{stats_file}"
    current_experiment = int(stats_file[stats_file.rfind("_") + 1:stats_file.rfind(".")])
    current_column = 5 + current_experiment
    packet_size = int(stats_file[stats_file.find("_") + 1:stats_file.rfind("_")])
    fer = foldername[foldername.find("_") + 1:]
    sheet = spreadsheet[f'Case {packet_size} FER {fer}']
    sheet["E1"] = packet_size
    header_bytes = 1 if packet_size <= 300 else 2
    profile = SigfoxProfile("UPLINK", "ACK ON ERROR", header_bytes)
    fragments = ceil((packet_size * 8) / (profile.UPLINK_MTU - profile.HEADER_LENGTH))
    sheet["E2"] = fragments
    windows = ceil(fragments / profile.WINDOW_SIZE)
    sheet["E3"] = windows
    output.write(f"-------------- RESULTS FOR {foldername}/{stats_file} --------------\n\n")
    with open(path, encoding='ISO-8859-1') as json_file:
        data = json.load(json_file)
    assert data['fragments']
    df1 = pd.read_json(str(json.dumps(data['fragments'], sort_keys=True)))
    df1_transposed = df1.T  # or df1.transpose()
    print(df1_transposed)
    df1_transposed.astype({"RULE_ID": str,
                           "W": str,
                           "FCN": str,
                           "data": str,  # or bytes?
                           "fragment_size": int,
                           "abort": bool,
                           "sending_start": float,
                           "sending_end": float,
                           "send_time": float,
                           "downlink_enable": bool,
                           "timeout": int,
                           "ack_received": bool,
                           "ack": str,  # or bytes?
                           "rssi": int,
                           "receiver_abort_received": bool,
                           "receiver_abort_message": str  # or bytes?
                           })
    df_nowait = df1_transposed[df1_transposed['downlink_enable'].isin([False])]
    output.write(f"Regular Fragments (nowait)\n"
                 # f"send_time =\n{df_nowait['send_time']}"
                 f"count: {df_nowait['send_time'].count()}\n"
                 f"sum: {df_nowait['send_time'].sum(axis=0, skipna=True)}\n"
                 f"mean: {df_nowait['send_time'].mean(axis=0, skipna=True)}\n"
                 f"std: {df_nowait['send_time'].std(axis=0, skipna=True)}\n"
                 f"\n")
    sheet.cell(row=6, column=current_column).value = df_nowait['send_time'].count()
    sheet.cell(row=8, column=current_column).value = df_nowait['send_time'].sum(axis=0, skipna=True)
    sheet.cell(row=9, column=current_column).value = df_nowait['send_time'].mean(axis=0, skipna=True)
    sheet.cell(row=10, column=current_column).value = df_nowait['send_time'].std(axis=0, skipna=True) if not isnan(
        df_nowait['send_time'].std(axis=0, skipna=True)) else 0
    df_wait = df1_transposed[df1_transposed['downlink_enable'].isin([True])]
    if len(df_wait[df_wait['RULE_ID'] == "00"]) != 0:
        df_all0 = df_wait[df_wait['FCN'].isin(['000'])]
        df_all1 = df_wait[df_wait['FCN'].isin(['111'])]
    else:
        df_all0 = df_wait[df_wait['FCN'].isin(['00000'])]
        df_all1 = df_wait[df_wait['FCN'].isin(['11111'])]
    output.write(f"Fragments - downlink requested - ALL 0\n"
                 f"data =\n{df_all0}\n"
                 f"send_time =\n{df_all0['send_time']}"
                 f"count: {df_all0['send_time'].count()}\n"
                 f"ul_errors: {df_all0[df_all0['ack_received'].isin([False])]['ack_received'].count()}\n"
                 f"all0_received: {df_all0[df_all0['ack_received'].isin([True])]['ack_received'].count()}\n"
                 f"sum: {df_all0['send_time'].sum(axis=0, skipna=True)}\n"
                 f"mean: {df_all0['send_time'].mean(axis=0, skipna=True)}\n"
                 f"std: {df_all0['send_time'].std(axis=0, skipna=True)}\n"
                 f"\n")
    sheet.cell(row=12, column=current_column).value = df_all0['send_time'].count()
    sheet.cell(row=13, column=current_column).value = df_all0[df_all0['ack_received'].isin([False])][
        'ack_received'].count()
    sheet.cell(row=15, column=current_column).value = df_all0[df_all0['ack_received'].isin([True])][
        'ack_received'].count()
    sheet.cell(row=16, column=current_column).value = df_all0['send_time'].sum(axis=0, skipna=True)
    sheet.cell(row=17, column=current_column).value = df_all0['send_time'].mean(axis=0, skipna=True) if not isnan(
        df_all0['send_time'].mean(axis=0, skipna=True)) else 0
    sheet.cell(row=18, column=current_column).value = df_all0['send_time'].std(axis=0, skipna=True) if not isnan(
        df_all0['send_time'].std(axis=0, skipna=True)) else 0
    output.write(f"Fragments - downlink requested - ALL 1\n"
                 f"data =\n{df_all1}\n"
                 f"send_time =\n{df_all1['send_time']}"
                 f"count: {df_all1['send_time'].count()}\n"
                 f"ul_errors: {df_all1[df_all1['ack_received'].isin([False])]['ack_received'].count()}\n"
                 f"all1_received: {df_all1[df_all1['ack_received'].isin([True])]['ack_received'].count()}\n"
                 f"sum: {df_all1['send_time'].sum(axis=0, skipna=True)}\n"
                 f"mean: {df_all1['send_time'].mean(axis=0, skipna=True)}\n"
                 f"std: {df_all1['send_time'].std(axis=0, skipna=True)}\n"
                 f"\n")
    sheet.cell(row=20, column=current_column).value = df_all1['send_time'].count()
    sheet.cell(row=21, column=current_column).value = df_all1[df_all1['ack_received'].isin([False])][
        'ack_received'].count()
    sheet.cell(row=23, column=current_column).value = df_all1[df_all1['ack_received'].isin([True])][
        'ack_received'].count()
    sheet.cell(row=24, column=current_column).value = df_all1['send_time'].sum(axis=0, skipna=True)
    sheet.cell(row=25, column=current_column).value = df_all1['send_time'].mean(axis=0, skipna=True)
    sheet.cell(row=26, column=current_column).value = df_all1['send_time'].std(axis=0, skipna=True) if not isnan(
        df_all1['send_time'].std(axis=0, skipna=True)) else 0
    # df1_transposed.to_excel('test_stats_2.2.xlsx', engine='xlsxwriter')
    output.write(
        f"Transmission Time (excluding code overhead): {df1_transposed['send_time'].sum(axis=0, skipna=True)}\n\n")
    sheet.cell(row=27, column=current_column).value = df1_transposed['send_time'].sum(axis=0, skipna=True)
    output.write(
        f"Transmission Time (of all the session): {data['total_transmission_time']}\n\n")
    sheet.cell(row=28, column=current_column).value = data['total_transmission_time']
    # is aborted?
    if df_all0[df_all0['abort'].isin([True])]['abort'].count() > 0:
        print("ABORT in df_all0")
        exit()
    elif df_all0[df_all0['receiver_abort_received'].isin([True])]['receiver_abort_received'].count() > 0:
        print("ReceiverAbort in df_all0")
        exit()
    elif df_all1[df_all1['abort'].isin([True])]['abort'].count() > 0:
        print("ABORT in df_all1")
        exit()
    elif df_all1[df_all1['receiver_abort_received'].isin([True])]['receiver_abort_received'].count() > 0:
        print("ReceiverAbort in df_all1")
        exit()
    output.write(
        f"tx_status_ok?: {data['tx_status_ok']}\n\n")
    sheet.cell(row=35, column=current_column).value = data["tx_status_ok"]


with open("output.txt", 'w', encoding='ISO-8859-1') as output_file:
    for foldername in ["results/ul_0",
                       "results/ul_10",
                       "results/ul_20",
                       "results/uldl_10",
                       "results/uldl_20"
                       ]:
        if foldername in ["results/ul_0", "results/ul_10", "results/ul_20"]:
            spreadsheet_name = 'results/Template Results UL.xlsx'
        else:
            spreadsheet_name = 'results/Template Results ULDL.xlsx'

        spreadsheet = openpyxl.load_workbook(filename=spreadsheet_name)
        output_file.write(f"============== FOLDER: {foldername} ==============\n\n")
        for filename in os.listdir(f'{foldername}'):
            print(f"Extracting data from {foldername}/{filename}")
            extract_data(foldername, filename, output_file, spreadsheet)
        spreadsheet.save(spreadsheet_name)
